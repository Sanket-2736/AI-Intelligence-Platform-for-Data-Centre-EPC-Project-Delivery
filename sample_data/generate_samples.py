"""
Sample Data Generator for EPC Intelligence Platform Demo

Generates realistic test data for all five agents:
- Schedule Excel with critical path and delays
- Shipments CSV with at-risk items and locations
- Specification PDFs (UPS system)
- Vendor submittal with intentional deviations
- All data ready for demo walkthrough
"""

import csv
import random
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Base paths
SAMPLE_DATA_DIR = Path(__file__).parent
SCHEDULE_FILE = SAMPLE_DATA_DIR / 'sample_schedule.xlsx'
SHIPMENTS_FILE = SAMPLE_DATA_DIR / 'sample_shipments.csv'
SPEC_FILE = SAMPLE_DATA_DIR / 'sample_ups_spec.txt'
SUBMITTAL_FILE = SAMPLE_DATA_DIR / 'sample_ups_submittal.txt'


def generate_schedule_excel(filename=SCHEDULE_FILE, num_tasks=80):
    """
    Generate realistic data centre construction schedule with critical path.
    
    Includes: Civil, MEP, Equipment Installation, Commissioning phases
    With: realistic delays, critical path tasks (0 float), and predecessors
    """
    print(f"Generating schedule: {filename}")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"
    
    # Headers
    headers = ['Task ID', 'Task Name', 'Phase', 'Start Date', 'Finish Date', 
               'Duration (days)', '% Complete', 'Predecessors', 'Resource', 
               'Baseline Start', 'Baseline Finish', 'Float (days)']
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Base dates
    project_start = datetime(2024, 6, 1)
    
    # Define phases with realistic tasks
    phases = {
        'Civil & Structure': [
            'Site Preparation & Excavation',
            'Foundation Work',
            'Structural Steel Erection',
            'Roof Installation',
            'Exterior Walls & Finish',
        ],
        'MEP Rough-In': [
            'MV Power Distribution Installation',
            'LV Electrical Panel Installation',
            'Cable Routing & Conduit',
            'Cooling Tower Installation',
            'Chiller System Rough-In',
            'AHU Installation',
            'Plumbing Rough-In',
        ],
        'Equipment Installation': [
            'UPS System Installation (Batt 1)',
            'UPS System Installation (Batt 2)',
            'Diesel Generator Set Installation',
            'Switchgear Installation',
            'PDU Installation & Testing',
            'Server Rack Installation',
            'Cooling Unit Final Installation',
        ],
        'Commissioning Prep': [
            'Power System Testing',
            'Cooling System Testing',
            'Network Infrastructure Test',
            'Fire Suppression Test',
            'Security System Test',
            'Final Inspections',
            'Punch List Resolution',
            'Handover Documentation',
        ],
    }
    
    row = 2
    task_id = 1
    phase_predecessors = {}
    
    for phase, tasks in phases.items():
        phase_start_offset = sum(len(t) for p, t in list(phases.items())[:list(phases.keys()).index(phase)]) * 8
        
        for task_idx, task_name in enumerate(tasks):
            tid = f"T{task_id:03d}"
            duration = random.randint(5, 15)
            task_start = project_start + timedelta(days=phase_start_offset + task_idx * 8)
            task_finish = task_start + timedelta(days=duration)
            baseline_start = task_start
            
            # Some tasks delayed
            if random.random() < 0.25:
                baseline_finish = task_finish
                task_finish = task_finish + timedelta(days=random.randint(2, 8))
                pct_complete = random.randint(60, 95)
                float_days = 0
            else:
                baseline_finish = task_finish
                pct_complete = random.randint(70, 100)
                float_days = random.randint(0, 5)
            
            # Critical path (0 float)
            if task_id % 12 == 0:
                float_days = 0
                pct_complete = random.randint(80, 100)
            
            # Predecessors
            predecessors = []
            if task_id > 1 and random.random() < 0.6:
                predecessors.append(f"T{task_id-1:03d}")
            
            ws.append([
                tid,
                task_name,
                phase,
                task_start.strftime('%Y-%m-%d'),
                task_finish.strftime('%Y-%m-%d'),
                duration,
                pct_complete,
                ','.join(predecessors),
                random.choice(['Contractor A', 'Contractor B', 'Electrical Team', 'MEP Team']),
                baseline_start.strftime('%Y-%m-%d'),
                baseline_finish.strftime('%Y-%m-%d'),
                float_days,
            ])
            
            task_id += 1
    
    # Auto-column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    print(f"✓ Schedule saved: {filename} ({task_id-2} tasks)")
    return filename



def generate_shipments_csv(filename=SHIPMENTS_FILE):
    """
    Generate 20 realistic equipment shipments with at-risk items and global origins.
    """
    print(f"Generating shipments: {filename}")
    
    # Origin locations: (country, lat, lng)
    origins = {
        'Germany': ('Frankfurt', 8.68, 50.11),
        'USA': ('Chicago', -87.63, 41.88),
        'Singapore': ('Singapore', 103.82, 1.35),
        'India': ('Mumbai', 72.88, 19.07),
        'China': ('Shenzhen', 114.06, 22.54),
    }
    
    # Equipment with suppliers
    equipment = [
        ('UPS System 500kVA', 'Eaton', 'Germany', 2500000),
        ('UPS System 750kVA', 'ABB', 'USA', 3200000),
        ('Diesel Generator Set 1MW', 'Cummins', 'Germany', 1800000),
        ('Diesel Generator Set 500kVA', 'BHEL', 'India', 900000),
        ('Diesel Generator Set 750kVA', 'MTU', 'USA', 1200000),
        ('Cooling Tower Unit 100TR', 'Thermax', 'India', 2000000),
        ('Cooling Tower Unit 150TR', 'Daikin', 'Singapore', 2500000),
        ('Centrifugal Chiller 500TR', 'Trane', 'USA', 5000000),
        ('Centrifugal Chiller 750TR', 'Carrier', 'USA', 6500000),
        ('AHU Unit 10000 CMH', 'Johnson Controls', 'Singapore', 800000),
        ('AHU Unit 15000 CMH', 'Tradewinds', 'India', 600000),
        ('11kV Switchgear Panel', 'Siemens', 'Germany', 3500000),
        ('11kV Switchgear Panel', 'ABB', 'Singapore', 3200000),
        ('MV Transformer 11/0.4kV', 'ABB', 'Germany', 2800000),
    ]
    
    today = datetime.now().date()
    
    rows = []
    for i, (eq_name, supplier, origin_country, cost) in enumerate(equipment * 2, 1):
        if i > 20:
            break
        
        origin_city, origin_lat, origin_lng = origins[origin_country]
        
        # Transit locations (various ports/cities)
        transit_locations = [
            origin_city,
            'Port of Rotterdam' if origin_country == 'Germany' else 'Port of Shanghai',
            'Suez Canal',
            'Port of Singapore',
            'Mumbai Port',
            'In Transit to Site',
        ]
        current = random.choice(transit_locations)
        
        # Realistic ETAs
        base_eta = today + timedelta(days=random.randint(10, 60))
        required = today + timedelta(days=random.randint(25, 90))
        
        # Create 3 at-risk, 2 delayed
        if i <= 3:
            # AT_RISK: buffer < 7 days
            buffer = random.randint(1, 6)
            required = base_eta + timedelta(days=buffer)
            status = 'at_risk'
        elif i <= 5:
            # DELAYED: ETA > Required
            base_eta = required + timedelta(days=random.randint(2, 10))
            status = 'delayed'
        else:
            status = 'in_transit'
        
        # Current location coordinates (rough path from origin to Mumbai site at 19.15, 72.87)
        if current == origin_city:
            curr_lat, curr_lng = origin_lat, origin_lng
        else:
            # Random point between origin and Mumbai
            curr_lat = origin_lat + (19.15 - origin_lat) * random.random()
            curr_lng = origin_lng + (72.87 - origin_lng) * random.random()
        
        rows.append({
            'equipment_name': f"{eq_name} #{i}",
            'supplier': supplier,
            'origin_country': origin_country,
            'current_location': current,
            'eta': base_eta.isoformat(),
            'required_on_site': required.isoformat(),
            'status': status,
            'lat': round(curr_lat, 4),
            'lng': round(curr_lng, 4),
            'po_number': f"PO-{2024000+i}",
            'cost_usd': cost,
        })
    
    # Write CSV
    with open(filename, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)
    
    print(f"✓ Shipments saved: {filename} ({len(rows)} items, 3 at-risk, 2 delayed)")
    return filename



def generate_sample_spec_text(filename=SPEC_FILE):
    """
    Generate realistic UPS system specification (3 pages).
    """
    print(f"Generating specification: {filename}")
    
    spec = """
================================================================================
                    HYPERSCALE DATA CENTRE SPECIFICATION
                        UNINTERRUPTIBLE POWER SUPPLY (UPS) SYSTEM
                                Version 2.1 - Final
================================================================================

1. SCOPE
--------

This specification defines the technical requirements for a redundant
Uninterruptible Power Supply (UPS) system for the Hyperscale Data Centre
Phase 1 project in Mumbai, India.

The UPS system shall provide continuous power conditioning and backup power
supply to all critical IT and facility infrastructure during:
- Utility mains failures or anomalies
- Planned maintenance of primary power systems
- Transition to backup generator sets

The system shall comply with TIA-942 Tier III redundancy standards and
Uptime Institute certification requirements.

2. TECHNICAL REQUIREMENTS
-------------------------

2.1 Power Rating
- Rated Capacity: 500 kVA (continuous 8-hour discharge)
- Module Configuration: 2x 250 kVA modules in parallel redundancy (2N)
- Input Voltage: 3-phase, 380V ±20%, 50 Hz ±5%
- Output Voltage: 3-phase, 380V ±5%, 50 Hz ±0.1%
- Frequency Range: 45-65 Hz input tolerance

2.2 Battery System
- Technology: VRLA (Valve-Regulated Lead-Acid) or Lithium-Ion
- Runtime: Minimum 10 minutes at full load (500 kVA)
- Discharge Time: ≤ 30 seconds to 90% load capacity
- Battery String: Fully monitored with cell-level balancing
- Charging Time: 24 hours to 100% SOC
- Operational Temperature: 15°C to 35°C
- Humidity: 20% to 80% (non-condensing)

2.3 Efficiency
- Overall System Efficiency: Minimum 96% (ECO mode)
- No-Load Loss: ≤ 150W per module
- Conversion Time: ≤ 4 milliseconds to battery power
- Noise Level: ≤ 75 dBA at 1 meter (6 ft)

2.4 Redundancy & Reliability
- Parallel Module Configuration: N+1 (one module failure acceptable)
- MTBF: ≥ 100,000 hours
- Automatic Load Transfer: Seamless, no transient spikes
- Hot-swap Battery Modules: Supported
- Preventive Maintenance: Possible without system shutdown

3. ENVIRONMENTAL REQUIREMENTS
-----------------------------

3.1 Mechanical
- Seismic Design: Zone 2, Response Spectrum per IS 1893:2016
- Vibration Isolation: Built-in or pedestal-mounted
- Acoustic Enclosure: Optional, ≤ 70 dBA with enclosure

3.2 Operating Environment
- Ambient Temperature: 0°C to 40°C
- Storage Temperature: -10°C to 60°C
- Relative Humidity: 20% to 80% (non-condensing)
- Altitude: Up to 2500 m
- Installation: Indoor, climate-controlled room

3.3 Electrical Safety
- EMI/RFI Compliance: EN 61000-6-2, CE Mark required
- Leakage Current: ≤ 3.5 mA per phase
- Grounding: Solid earth (≤ 5 ohms)

4. TESTING REQUIREMENTS
-----------------------

4.1 Factory Acceptance Test (FAT)
- Full load test for 30 minutes
- Battery discharge to 90% DOD
- Harmonic distortion measurement (THD ≤ 5%)
- Load transfer testing
- Documentation: Test report + video

4.2 Site Acceptance Test (SAT)
- Installation verification
- Electrical continuity & polarity check
- Battery capacity verification (80-minute runtime test)
- N+1 redundancy switchover simulation
- Monitoring system verification
- Final commissioning sign-off

4.3 Post-Installation
- Trending data collection for 72 hours
- Load profile analysis
- Battery health assessment every 12 months

================================================================================
APPENDICES
A - Electrical Drawings (attached separately)
B - Equipment Datasheet
C - Installation & Maintenance Manual
================================================================================
    """
    
    with open(filename, 'w') as f:
        f.write(spec)
    
    print(f"✓ Specification saved: {filename}")
    return filename



def generate_sample_submittal_text(filename=SUBMITTAL_FILE):
    """
    Generate UPS vendor submittal with 3-4 intentional deviations from spec.
    """
    print(f"Generating submittal: {filename}")
    
    submittal = """
================================================================================
                        VENDOR SUBMITTAL - UPS SYSTEM
                    Eaton 9PXM 500 kVA UPS with Battery
                              Dated: 2024-05-15
================================================================================

VENDOR: Eaton Corporation
MODEL: 9PXM 500 UPS System
PART NUMBER: 9PXMEM0500ADAD00

1. POWER RATINGS
----------------
- Rated Capacity: 500 kVA (continuous)
- Module Configuration: Single 500 kVA unit (standard factory config)
- Input: 3-phase 380V, 50 Hz ±10% (broader tolerance than spec)
- Output: 3-phase 380V ±8%, 50 Hz ±0.2%
- Efficiency (ECO Mode): 95% (DEVIATION: Spec requires 96%)

2. BATTERY SPECIFICATIONS
-------------------------
- Type: VRLA Lead-Acid Battery Module
- Capacity: 80 Ah per string
- Battery Runtime: 8 MINUTES at full 500 kVA load (DEVIATION: Spec requires 10 min)
- Discharge Profile: Fast discharge, optimized for short-duration bridging
- Temperature Range: 15°C to 30°C
- Charging Time: 24 hours standard

3. ENVIRONMENTAL & SEISMIC
---------------------------
- Operating Temperature: 0°C to 40°C (compliant)
- Humidity: 20% to 80% (compliant)
- Altitude: Up to 2500 m (compliant)
- SEISMIC RATING: Not specified in this submittal (DEVIATION: Spec mandates Zone 2 IS 1893)
- Installation: Indoor, climate-controlled only

4. SAFETY & COMPLIANCE
----------------------
- CE Mark: Yes
- EMI/RFI: EN 61000-6-2 Compliant
- Leakage Current: 3.8 mA per phase (MINOR: Spec limit is 3.5 mA)
- Grounding: Standard earthing required (≤ 5 ohms)

5. REDUNDANCY NOTE
------------------
- This is a single-unit configuration
- Customer must specify if 2N redundancy required (order 2x units)
- Each unit is independent with separate battery

6. TESTING CERTIFICATE
----------------------
- Factory test performed: 2024-04-10
- Test Duration: 20 minutes (less than spec's required 30 min)
- Load Transfer Time: 5 ms
- Harmonic Content: 4.2% THD (compliant)

7. DOCUMENTATION PROVIDED
--------------------------
✓ Equipment Datasheet (attached)
✓ Installation Manual
✓ Maintenance Schedule
✗ Electrical Drawings (to be provided at FAT)
✗ Full Test Protocol (standard factory test only)

8. DELIVERY & INSTALLATION
---------------------------
- Lead Time: 18 weeks FOB Germany
- Shipping: CFR (Cost and Freight) to Mumbai
- Installation Support: Optional paid service
- Warranty: 2 years parts & labor

VENDOR NOTES:
This is our standard 9PXM configuration. Non-standard battery runtime or
seismic requirements will require engineering review (4-week delay).
Efficiency at 95% is industry-standard for this product class.

Submitted by: Eaton Sales Team, Germany
Date: 2024-05-15

================================================================================
    """
    
    with open(filename, 'w') as f:
        f.write(submittal)
    
    print(f"✓ Submittal saved: {filename} (with 4 intentional deviations)")
    return filename



if __name__ == '__main__':
    """
    Generate all sample data for demo.
    """
    print("\n" + "="*80)
    print("EPC INTELLIGENCE PLATFORM - SAMPLE DATA GENERATOR")
    print("="*80 + "\n")
    
    SAMPLE_DATA_DIR.mkdir(parents=True, exist_ok=True)
    
    # Generate all samples
    generate_schedule_excel()
    generate_shipments_csv()
    generate_sample_spec_text()
    generate_sample_submittal_text()
    
    print("\n" + "="*80)
    print("✓ All sample data generated successfully!")
    print("="*80)
    print(f"\nLocation: {SAMPLE_DATA_DIR}")
    print("\nReady for demo:")
    print("  1. Upload sample_schedule.xlsx to Schedule Risk Agent")
    print("  2. Upload sample_shipments.csv to Supply Chain Agent")
    print("  3. Upload sample_ups_spec.txt + sample_ups_submittal.txt to Compliance Agent")
    print("  4. Upload sample_ups_spec.txt to RFI Agent for Q&A\n")
